# -*- coding: utf-8 -*-

from datetime import datetime
import unicodedata
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from charset_normalizer import md__mypyc ## para el ejecutable
# C:\Users\marti\OneDrive\Escritorio
ruta_excel = 'C:\\Users\\marti\\OneDrive\\Escritorio\\FvCata.xlsx'

class Producto:
    def __init__(self, codigo, link, nombre = '//*[@id="root-app"]/div/div[2]/div/div/div/div/div/a/div/div/h1',
                 primero = ['/html/body/main/div/div[3]/div/div/div/form[1]/div[1]/div[1]/div/div/span[1]/span/span[2]','/html/body/main/div/div[3]/div/div/div/form[1]/div[4]/div/div/div/div/div/div/div/div/a/span'],
                 segundo = ['/html/body/main/div/div[3]/div/div/div/form[2]/div[1]/div[1]/div/div/span/span/span[2]','/html/body/main/div/div[3]/div/div/div/form[2]/div[4]/div/div/div/div/div/div/div/div/a/span']):
        self.codigo = codigo
        self.link = link
        self.nombre = nombre
        self.primero = primero
        self.segundo = segundo
class SeniorExcel:
    def __init__(self, excel):
        self.excel = excel

    def cargar_productos(self,hoja_destino):
        lista_productos = []
        print("cargo urls")
        for row in self.excel[hoja_destino].iter_rows(min_row=3, max_row=779, min_col=1, max_col=1):
            if row[0].value is not None:
                print(row[0].value,row[0].offset(column=7).hyperlink.target)
                lista_productos.append(Producto(codigo=row[0].value,link=row[0].offset(column=7).hyperlink.target))
                continue
            break
            
        print("Me fui")
        return lista_productos
    def escribirExcel(self,lista_a_cargar,hoja_destino):
    # Cargar el archivo Excel

        print("Estoy en la hoja")
        
        for row in self.excel[hoja_destino].iter_rows(min_row=3, max_row=779, min_col=1, max_col=1):
            for producto_cargado in lista_a_cargar:
                if producto_cargado[0] == (row[0].value):
                    print("Coincidencia encontrada. Sobrescribiendo...")
                    # Sobrescribir los datos en la fila correspondiente
                    row[0].offset(column=1).value = producto_cargado[1]
                    row[0].offset(column=2).value = producto_cargado[2]
                    row[0].offset(column=3).value = producto_cargado[3]
                    row[0].offset(column=4).value = producto_cargado[4]
                    row[0].offset(column=5).value = producto_cargado[5]
                    row[0].offset(column=6).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    break
        
        # Guardar el archivo Excel
        self.excel.save(ruta_excel)
        


    
        

# urls_fv = [
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-ducha-y-banera-fv-puelo-0106b5-color-cromo/p/MLA15998519/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/fv-coty-0106d9-color-cromo/p/MLA15842508/s',
#     Producto()'https://www.mercadolibre.com.ar/griferia-fv-para-ducha-monocomando-108m4-compacta/p/MLA22225320/s',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-ducha-simple-fv-arizona-plus-cromo-0109b1p-cr/p/MLA22315853/s?pdp_filters=item_id:MLA1366802411',
#     Producto()'https://www.mercadolibre.com.ar/juego-griferia-ducha-fv-exterior-arizona-plus-113b1p-cr/p/MLA22393870/s?pdp_filters=item_id:MLA1366741583',
#     Producto()'https://www.mercadolibre.com.ar/brazo-para-ducha-fv-duchamatic-cromo-duchador-011902/p/MLA22302114/s?pdp_filters=category:MLA4334',
#     Producto()'https://www.mercadolibre.com.ar/ducha-cbrazo-fv-12017-california-anticalcarea/p/MLA22225378/s?pdp_filters=category:MLA1613',
#     Producto()'https://www.mercadolibre.com.ar/toallero-de-aro-fv-libby-016239-bano/p/MLA21902981/s?pdp_filters=category:MLA31034',
#     Producto()'https://www.mercadolibre.com.ar/toallero-corto-barral-fv-arizona-0163b1/p/MLA21819034/s',
#     Producto()'https://www.mercadolibre.com.ar/fv-puelo-018101b5-color-cromo/p/MLA15581140/s?pdp_filters=category:MLA1574',
#     Producto()'https://www.mercadolibre.com.ar/fv-libby-01810239-color-cromo/p/MLA15344326/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/fv-dominic-new-01810285n-color-cromo/p/MLA15383207/s?pdp_filters=category:MLA4334',
#     Producto()'https://www.mercadolibre.com.ar/fv-87-temple-01810287-color-cromo/p/MLA15551684/s',
#     Producto()'https://www.mercadolibre.com.ar/fv-puelo-018102b5-color-cromo/p/MLA15404750/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/fv-coty-018102d9-cromo/p/MLA15379701/s',
#     Producto()'https://www.mercadolibre.com.ar/fv-epuyen-018102l2-color-negra-mate/p/MLA15448586/s?pdp_filters=category:MLA1574',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-bano-fv-libby-018139-color-cromo-y-acabado-brillante/p/MLA15383246/s',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-arizona-0181b1-color-cromo/p/MLA15383244/s?pdp_filters=category:MLA1500',
#     Producto()'https://www.mercadolibre.com.ar/fv-puelo-0181b5-color-cromo-y-acabado-cromado/p/MLA15462375/s',
#     Producto()'https://www.mercadolibre.com.ar/fv-coty-0181d9-color-cromo/p/MLA15379709/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/fv-epuyen-0181l2-color-negra-y-acabado-mate/p/MLA15551675/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-jana-m1-0181m1-cromo/p/MLA21719931/s?pdp_filters=category:MLA412368',
#     Producto()'https://www.mercadolibre.com.ar/fv-compact-monocomando-lavatorio-con-desague-oferta/p/MLA19783121/s?pdp_filters=category:MLA1574',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-libby-018939-color-cromo/p/MLA15398002/s',
#     Producto()'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-temple-018987-color-cromo/p/MLA15912514/s',
#     Producto()'https://www.mercadolibre.com.ar/fv-arizona-0189b1-color-cromo/p/MLA15581138/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-puelo-0189b5-color-cromo-y-acabado-cromado/p/MLA15460630/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-coty-0189d9-color-cromo/p/MLA15836455/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0189l2-color-negra-y-acabado-mate/p/MLA15456769/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-jana-m1-0189m1-cromo/p/MLA22847544/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-compact-monocomando-bidet-con-transferencia-0189m4/p/MLA19753497/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-libby-020339-color-cromo-y-acabado-cromado/p/MLA15551672/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-alerce-0203d7-color-cromo/p/MLA16023790/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-libby-monocomando-020639-color-cromo/p/MLA15735567/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-new-020685n-color-cromo-y-acabado-cromado/p/MLA15400029/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0206l2-color-negra-y-mate/p/MLA15400024/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-allegro-020715-color-cromo/p/MLA15592911/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-malena-020716-color-cromo/p/MLA16023866/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-california-020717-color-cromo/p/MLA15592899/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-libby-020739-color-cromo/p/MLA15344324/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-arizona-plus-0207b1p-color-cromo/p/MLA15581350/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-newport-plus-0207b2p-color-cromo-y-acabado-cromado/p/MLA15398000/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-pampa-0207b6-color-cromo-y-acabado-cromado/p/MLA15402622/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-triades-0207c3-color-cromo/p/MLA15551682/s?pdp_filters=category:MLA1613',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-radal-0207c7-cromo/p/MLA15592903/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-alerce-0207d7-color-cromo-y-acabado-cromado/p/MLA15756273/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bacha-fv-epuyen-0207l2-color-cromo/p/MLA15781046/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-bidet-fv-malena-029516-color-cromo/p/MLA15754384/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-california-029517-color-cromo/p/MLA15735460/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-libby-029539-color-cromo/p/MLA15400027/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-arizona-plus-0295b1p-color-cromo/p/MLA15714990/s?pdp_filters=category:MLA417758',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-newport-plus-0295b2p-color-cromo/p/MLA15592905/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-pampa-0295b6-color-cromo-y-acabado-cromado/p/MLA15462457/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-radal-0295c7-color-cromo/p/MLA15592907/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-alerce-0295d7-color-cromo/p/MLA15735462/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0295l2-color-cromo/p/MLA15753928/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/duchador-monocomando-fv-arizona-0310b1-griferia-exterior/p/MLA21815193/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-banera-monocomando-epuyen-negro-mate-0310l2-ng/p/MLA21671879/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-ducha-fv-310m4-compacta-monocomando/p/MLA22982682/s?pdp_filters=category:MLA1613',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b2p-newport-plus-0403b2p-cromo-acabado-cromado/p/MLA9305647/s?pdp_filters=category:MLA53260',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-39-libby-04060339-cromo/p/MLA15183206/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b1-arizona-0406b1-cromo/p/MLA7984947/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b5-puelo-0406b5-cromo-acabado-cromado/p/MLA9305069/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b1p-arizona-plus-040902b1p-cromo/p/MLA9343806/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-15-allegro-040915-cromo/p/MLA9305338/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-b1p-arizona-plus-0409b1p-cromo/p/MLA9337391/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-90-swing-04110190-cromo/p/MLA8064798/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b1-arizona-041101b1-cromo/p/MLA8283526/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b2-newport-041101b2-cromo/p/MLA8338628/s?pdp_filters=category:MLA53260',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-compacta-41101m4-cocina-mesada-monocomando/p/MLA19770131/s?pdp_filters=category:MLA53260',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b1-arizona-041102b1-cromo-acabado-cromado/p/MLA7984961/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b2-newport-041102b2-cromo/p/MLA8018650/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b1-arizona-041103b1-cromo/p/MLA7984945/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-24-kansas-04110424-cromo/p/MLA7978914/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-27-alabama-04110427-cromo-acabado-cromado/p/MLA7984949/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-39-libby-04110439-cromo/p/MLA8050800/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-39-libby-04110439e-cromo/p/MLA15957472/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-87-temple-04110487-acero/p/MLA15152437/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-d8-aromo-041104d8-cromo/p/MLA9467786/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-l2-epuyen-041104l2-cromo/p/MLA9467721/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-l2-epuyen-041104l2-negro-mate/p/MLA15761967/s?pdp_filters=category:MLA1574',
#    Producto(codigo ='',link='') 'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-87-temple-041187-cromo/p/MLA9202729/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-87-temple-04120187-extensible-cromo-acabado-cromado/p/MLA7978918/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-87-temple-041287-cromo-cromado/p/MLA7984943/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-h5-lenga-0412h5-cromo/p/MLA9928500/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b1p-arizona-plus-0413b1p-cromo-acabado-cromado/p/MLA8338630/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b2p-newport-plus-0413b2p-cromo/p/MLA7984957/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b6-pampa-0413b6-cromo-acabado-cromado/p/MLA9295275/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-c7-radal-0413c7-cromo/p/MLA9696532/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-g5-obera-0414g5-cromo/p/MLA15762090/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-15-allegro-041615-cromo/p/MLA9322049/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-17-california-041617-cromo/p/MLA9293463/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-15-allegro-042015-cromo-acabado-cromado/p/MLA10026388/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-15-allegro-042115-cromo-acabado-brillante/p/MLA10015798/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b1-arizona-0423b1-cromo/p/MLA8338657/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b5-puelo-0423b5-cromo-acabado-cromado/p/MLA9311597/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-cocina-fv-compacta-cromo-monocomando-0423m4-cromo/p/MLA21016908/s?pdp_filters=item_id:MLA1308307621',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-15-allegro-042515-cromo-acabado-cromado/p/MLA11663778/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-d7-alerce-0428d7-cromo-acabado-cromado/p/MLA8050792/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-d7l-alerce-lever-0428d7l-cromo/p/MLA13796874/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-libby-banera-ducha-con-transferencia-10339/p/MLA21843566/s?pdp_filters=item_id:MLA1371070705',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-bano-fv-pampa-ducha-con-transferencia-0103b6-color-cromo/p/MLA25903508/s?pdp_filters=item_id:MLA1474435978',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-alerce-ducha-embutida-cierre-ceramico-0103d7/p/MLA21801086/s?pdp_filters=item_id:MLA1451879086',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0103l2-color-cromo/p/MLA15837220/s?pdp_filters=item_id:MLA1471984354',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-new-010685n-cromo/p/MLA15390241/s?pdp_filters=item_id:MLA1383184291',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0106l2-cromo-cromado/p/MLA15383203/s?pdp_filters=item_id:MLA1471930342',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0106l2-color-negra-y-acabado-mate/p/MLA15383202/s?pdp_filters=item_id:MLA1383309853',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-libby-monocomando-para-ducha-sin-transferencia-010839/p/MLA22366925/s?pdp_filters=item_id:MLA1383378801',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/ducha-fv-200-mm-articulada-metalica-redonda-0126mr0-20-cr/p/MLA22294101/s?pdp_filters=item_id:MLA1472006072',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/ducha-articulada-fv-126pro-20cm-plastica-redonda-p/p/MLA22225396/s?pdp_filters=item_id:MLA1471891150',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-de-aro-fv-libby-016239-bano/p/MLA21902981/s?pdp_filters=category:MLA31034',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-bano-fv-libby-toallero-barral-cromo-16439/p/MLA24548703/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/portarrollo-fv-epuyen-cromo-167l2/p/MLA22653627/s?pdp_filters=item_id:MLA1435322352',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-arizona-porta-vaso-cromo-169b1/p/MLA24159431/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-arizona-porta-vaso-cromo-169b1/p/MLA24159431/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-libby-01810239-color-cromo/p/MLA15344326/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-aromo-018102d8-color-cromo/p/MLA15551680/s?pdp_filters=item_id:MLA1471216806',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-new-018185n-cromo-cromado/p/MLA15456721/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-aromo-0181d8-color-cromo/p/MLA15815622/s?pdp_filters=item_id:MLA1383162137',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-new-018985n-color-cromo/p/MLA15581242/s?pdp_filters=item_id:MLA1383113173',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-triades-0203c3-color-cromo-y-acabado-cromado/p/MLA15551677/s?pdp_filters=item_id:MLA1473236816',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-margot-lever-020762l-color-cromo-y-acabado-cromado/p/MLA15718855/s?pdp_filters=item_id:MLA1383368127',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-denisse-lever-020764l-color-cromo/p/MLA16174749/s?pdp_filters=item_id:MLA1471989384',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-puelo-018101b5-color-cromo/p/MLA15581140/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-lever-0207r85l-color-cromo/p/MLA15398040/s?pdp_filters=item_id:MLA1471977862',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-margot-lever-029562l-color-cromo/p/MLA15591512/s?pdp_filters=item_id:MLA1383475015',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-triades-0295c3-color-cromo/p/MLA15810248/s?pdp_filters=item_id:MLA1383095941',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-lever-0295r85l-color-cromo/p/MLA15735569/s?pdp_filters=item_id:MLA1471886608',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/tapa-tecla-fv-doble-descarga-para-valvula-inodoro-036804-cr/p/MLA24156956/s?pdp_filters=seller_id:142125695',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-b1p-arizona-plus-0403b1p-cromo/p/MLA9895144/s?pdp_filters=item_id:MLA1471660354',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-39-libby-04060339-cromo/p/MLA15183206/s?pdp_filters=item_id:MLA1471660354',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-87-temple-04110487-acero/p/MLA15152437/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-16-malena-041616-cromo/p/MLA8050790/s?pdp_filters=item_id:MLA1472601400',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-62l-margot-lever-041662l-cromo/p/MLA9295279/s?pdp_filters=item_id:MLA1383356069',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-64-denisse-041664-cromo/p/MLA8050788/s?pdp_filters=item_id:MLA1383768659',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-fv-39-libby-042839-cromo-acabado-cromado/p/MLA8064606/s?pdp_filters=item_id:MLA1472200134',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/llave-de-gas-sin-campana-aprobada-mh-fv-12-081001-13-c/p/MLA25390106/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-b5-puelo-041104b5-cromo-acabado-cromado/p/MLA7974067/s?pdp_filters=category:MLA1574'
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-ducha-banera-fv-temple-monocomando-01060287-cr/p/MLA22288746/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-87-temple-018187-cromo/p/MLA15828533/s?pdp_filters=item_id:MLA1388030713',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-alesia-lever-020749l-color-cromo/p/MLA15816488/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-chalten-0207h4-color-cromo/p/MLA15904932/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-alesia-lever-029549l-color-cromo/p/MLA15735575/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-bano-fv-bidet-cipres-295n2-cromo-cierre-ceramico/p/MLA20005213/s?pdp_filters=category:MLA4334',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/rejilla-desague-lineal-fv-vista-reversible-60-cm-0351026/p/MLA26654255/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/tapa-y-tecla-para-valvula-descarga-de-inodoro-fv-036802/p/MLA27610955/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/valvula-fv-slim-para-descarga-de-inodoros-037801-bano/p/MLA27652233/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/tecla-boton-slim-extra-chata-descarga-inodoro-fv-cromada/p/MLA27615243/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-87-temple-04110487-cromo/p/MLA7991448/s?product_trigger_id=MLA15152437&quantity=1',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-96-augusta-041296-extensible-cromo/p/MLA8030437/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-ducha-banera-fv-temple-monocomando-01060287-cr/p/MLA22288746/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/portavaso-portacepillo-fv-epuyen-negro-mate-0169l2-ng/p/MLA24305402/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-coty-018102d9-cromo/p/MLA15379701/s',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-monocomando-ducha-banera-libby-010639-cromo/p/MLA22209424/s?pdp_filters=item_id:MLA1395582629',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0106l2-color-negra-y-acabado-mate/p/MLA15383202/s?pdp_filters=item_id:MLA1383309853',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/flor-de-ducha-articulada-cuadrada-25-cm-fv-0126-pco-25/p/MLA22426443/s?pdp_filters=item_id:MLA1395582703',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-brazo-de-ducha-vertical-12cm-cromo/p/MLA22302115/s?pdp_filters=item_id:MLA1395608707',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-aro-fv-alesia-cromo-016249-cr/p/MLA21790047/s?pdp_filters=item_id:MLA1395582895',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-fv-chalten-164h4-cr-toallero-barral/p/MLA22982332/s?pdp_filters=item_id:MLA1563607204',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-barral-fv-cipres-0164n2-cr/p/MLA21790045/s?pdp_filters=item_id:MLA1564050698',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-fv-percha-para-bano-alesia-016649/p/MLA22331583/s?pdp_filters=item_id:MLA1564063558',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-portarrollo-fv-dominic-cromo-167r85-cr/p/MLA23431908/s?pdp_filters=item_id:MLA1564089612',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-libby-jabonera-cromo-16839/p/MLA25765538/s?pdp_filters=item_id:MLA1566882892',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/jabonera-fv-temple-cromo-16887-accesorio-bano-crom/p/MLA25765544/s?pdp_filters=item_id:MLA1567125446',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/jabonera-fv-vidrio-accesorios-bano-epuyen-0168l2-cromo/p/MLA26694843/s?pdp_filters=item_id:MLA1388671889',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-jabonera-epuyen-negro-mate-p/p/MLA26172772/s?pdp_filters=item_id:MLA1385288903',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/jabonera-fv-metal-accesorios-bano-dominic-r-0168r85m-cromo/p/MLA27824374/s?pdp_filters=item_id:MLA1395633867',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/portacepillos-fv-temple-16987-accesorio-bano-cromo/p/MLA24147739/s?pdp_filters=item_id:MLA1566856836',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-bano-fv-dominic-porta-vaso-cromo-16985/p/MLA24147719/s?pdp_filters=item_id:MLA1566856910',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-lavatorio-fv-epuyen-pico-bajo-cromo-acabado-cromado/p/MLA15551674/s?pdp_filters=item_id:MLA1566870530',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0189l2-color-cromo-y-acabado-cromado/p/MLA15456770/s?pdp_filters=item_id:MLA1566857540',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-alesia-020349/p/MLA15998455/s?pdp_filters=item_id:MLA1566870446',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-chalten-0203h4-color-cromo/p/MLA15986900/s?pdp_filters=item_id:MLA1566935546',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-bidet-fv-chalten-0295h4-color-chrome/p/MLA15846382/s?pdp_filters=item_id:MLA1566896082',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-epuyen-0310l2-color-cromo-y-acabado-cromado/p/MLA15462373/s?pdp_filters=item_id:MLA1572567640',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-monocomando-fv-jana-0106m1-ducha-banera-acabado-cromado-color-plateado/p/MLA27945475/s?pdp_filters=item_id:MLA1553897272',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-pampa-bano-ducha-embutir-s-transferencia-109b6-acabado-cromado-color-plateado/p/MLA25365324/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/flor-ducha-lluvia-fv-012604-b-antisarro-autolimpiable/p/MLA22732125/s?pdp_filters=item_id:MLA1617862324',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-temple-toallero-aro-cromo-16287/p/MLA22641646/s?pdp_filters=item_id:MLA1398645431',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-corto-fv-accesorio-bano-linea-epuyen-crom-163l2-cr/p/MLA22641645/s?pdp_filters=item_id:MLA1446298440',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-barral-fv-temple-cromo-016487-cr/p/MLA21827423/s?pdp_filters=item_id:MLA1606661166',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/toallero-barral-fv-accesorios-bano-dominic-r-0164r85-cromo/p/MLA22803282/s?pdp_filters=item_id:MLA1377216919',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-bano-porta-rollo-fv-libby-cromo-16739/p/MLA22022916/s?pdp_filters=item_id:MLA1424526772',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-temple-portarrollo-cromo-16787/p/MLA22809748/s?pdp_filters=item_id:MLA1607094436',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/porta-rollo-fv-arizona-167b1-accesorios-bano-cromo/p/MLA22515908/s?pdp_filters=category:MLA1613',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorios-fv-chalten-167h4-cr-portarrollo-p/p/MLA22957959/s?pdp_filters=item_id:MLA1398645133',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/accesorio-bano-fv-arizona-jabonera-cromo-168b1/p/MLA26374477/s?pdp_filters=item_id:MLA1553868484',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-de-lavatorio-pared-fv-cipres-cierre-ceramico-203n2/p/MLA26042596/s?pdp_filters=item_id:MLA1383470799',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/fv-dominic-new-020685n-color-cromo-y-acabado-cromado/p/MLA15400029/s?pdp_filters=category:MLA412368',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/valvula-fv-angular-de-12-m-m-026905-cr-macho-a-macho/p/MLA27234333/s?pdp_filters=category%3AMLA1500',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-bidet-fv-arrayan-0295h6-acabado-brillante-color-cromo/p/MLA28220210/s?pdp_filters=item_id:MLA1397772363',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/valvula-automatica-descarga-inodoro-fv-36801-para-bano/p/MLA28082261/s?pdp_filters=item_id:MLA1395426579',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-tapa-tecla-valvula-descarga-inodoro-bano-036802/p/MLA28698508/s?pdp_filters=item_id:MLA1593327322',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/griferia-fv-arizona-lavadero-lavatorio-pared-embutir-401b1p/p/MLA21415040/s?pdp_filters=category:MLA1574',
#     Producto(codigo ='',link='')'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-fv-39-libby-04060339-cromo/p/MLA15183206/s'
# ]

# urls_ferrum = [
#     Producto(codigo ='1001-Ferrum',link='https://www.mercadolibre.com.ar/bacha-ferrum-apoyar-country-baja-blanca-59x38x13cml16kf/p/MLA23132534/s?pdp_filters=item_id:MLA1423143978'),
#     Producto(codigo ='1002-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-ferrum-persis-de-apoyo-negra-redonda-l17kfn/p/MLA22922607/s?pdp_filters=item_id:MLA1368570371',
#     Producto(codigo ='1003-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-de-apoyar-ferrum-tori-cuenco-lk060-blanco-34cm-x-40cm-x-338cm-145cm-de-alto/p/MLA10459327/s?pdp_filters=category:MLA412391',
#     Producto(codigo ='1004-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-bajo-mesada-lavatorio-ferrum-esquel-blanco-lebf-b/p/MLA21808230/s',
#     Producto(codigo ='1005-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-sobre-mesada-ferrum-esquel-lesf-blanco-380mm-x-480mm-205mm-de-alto/p/MLA10663694/s?pdp_filters=item_id:MLA1423106534',
#     Producto(codigo ='1006-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-apoyo-rectangular-country-ferrum/p/MLA20730678/s?pdp_filters=item_id:MLA1421039566',
#     Producto(codigo ='1007-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-bajo-mesada-ferrum-cuadra-lfnf-blanco-358cm-x-452cm/p/MLA9588915/s?pdp_filters=category:MLA411920',
#     Producto(codigo ='1008-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-bano-dongio-semiencastrada-ferrum-tori-lks10-b-apoyo/p/MLA21209983/s?pdp_filters=item_id:MLA1423108904',
#     Producto(codigo ='1009-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-bajo-mesada-ferrum-cadria-lnsf-blanco/p/MLA9943370/s',
#     Producto(codigo ='1010-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-ferrum-persis-lwpf-blanco-164mm-de-alto-355mm-de-diametro/p/MLA9282980/s?pdp_filters=item_id:MLA1399462196',
#     Producto(codigo ='1011-Ferrum',link='')'https://www.mercadolibre.com.ar/bacha-de-bano-ferrum-persis-lwpf-negro-164mm-de-alto-355mm-de-diametro/p/MLA9569102/s?pdp_filters=item_id:MLA1399197224',
#     Producto(codigo ='1012-Ferrum',link='')'https://www.mercadolibre.com.ar/asiento-para-inodoro-traful-blanco-ferrum-universal-tdxp/p/MLA22298856/s?pdp_filters=item_id:MLA1381050921',
#     Producto(codigo ='1013-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-trento-blanco-ferrum/p/MLA20030378/s',
#     Producto(codigo ='1014-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-para-inodoro-ferrum-trento-blanco-texc/p/MLA24001215/s?pdp_filters=item_id:MLA1455051682',
#     Producto(codigo ='1015-Ferrum',link='')'https://www.mercadolibre.com.ar/inodoro-ferrum-trento-tapa-texus-cierre-suave-herraje-cromo/p/MLA24080416/s?pdp_filters=item_id:MLA1455013298',
#     Producto(codigo ='1016-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-inodoro-ferrum-traful-andina-florencia-tfw/p/MLA19958827/s?pdp_filters=item_id:MLA1441828088',
#     Producto(codigo ='1017-Ferrum',link='')'https://www.mercadolibre.com.ar/ferrum-florenciaatuel-tapa-para-inodoros-herraje-nylon-tfxb/p/MLA19981770/s?pdp_filters=item_id:MLA1375463925',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-inodoro-dorica-florencia-ferrum-tfxcb-cromo/p/MLA19945251/s?pdp_filters=item_id:MLA1441827780',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/asiento-para-inodoro-blanco-ferrum-bari-tkwps/p/MLA20670056/s?pdp_filters=item_id:MLA1425197998',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/ferrum-bari-tapa-para-inodoros-herraje-nylon-blanco-tkxm/p/MLA23022584/s?pdp_filters=item_id:MLA1375487765',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-inodoro-ferrum-bari-tkxmc-herrajes-cromo-blanco/p/MLA22836364/s?pdp_filters=item_id:MLA1441866764',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-inodoro-ferrum-andina-plastico/p/MLA20724112/s?pdp_filters=category:MLA411920',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/asiento-para-inodoro-ferrum-traful-tsx-de-hdf-con-forma-redonda-blanco-liso/p/MLA19816134/s',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/ferrum-veneto-tapa-de-asiento-pinod-cierre-suave-ttuxs-bco/p/MLA21088698/s?pdp_filters=category:MLA411920',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-asiento-inodoro-ferrum-murano-tux-madera-laqueada/p/MLA23437962/s?pdp_filters=item_id:MLA1454375906',
#     Producto(codigo ='1000-Ferrum',link='')'https://www.mercadolibre.com.ar/tapa-inodoro-blanco-herraje-metal-ferrum-murano/p/MLA20030395/s?pdp_filters=item_id:MLA1282631775'
# ]

# urls_peirano = [
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/kit-accesorios-bano-peirano-black-linea-10000-5-piezas-color-negro/p/MLA32431312/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-monocomando-cocina-peirano-fabric-gold-oro-flexible-color-dorado/p/MLA20028410/s?pdp_filters=category:MLA1574'
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-mesada-peirano-renacer-20-102-cromo-cromado-monocomando/p/MLA9272673/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-marbella-20-121-cromo-acabado-cromado/p/MLA8372565/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-mesada-peirano-betis-20-134-cromo-cromado-monocomando/p/MLA9235213/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-santander-20-135-plateado-acabado-cromado/p/MLA9323226/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-cuina-20-140-extensible-cromo-acabado-cromado/p/MLA8018648/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-monocomando-peirano-toledo-20-150-color-cromo/p/MLA8848674/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-teruel-20-154-cromado/p/MLA15171356/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-dique-20-160-cromo-acabado-cromado/p/MLA9343917/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-monocomando-cocina-mesada-peirano-adra-20-162-cromo/p/MLA15727517/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-colors-20-17-azul-acabado-cromado/p/MLA13031815/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-colors-20-17-gris-acabado-cromado/p/MLA13031817/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-colors-20-17-negro-acabado-cromado/p/MLA13031816/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-colors-20-17-rojo-acabado-cromado/p/MLA13031818/s?product_trigger_id=MLA13031817&quantity=1',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-monocomando-black-20-174-peirano-colors/p/MLA20731376/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-vera-colors-20-180-acabado-cromado/p/MLA17874466/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-monocomando-peirano-black-velvet-20-190-color-negro/p/MLA15187004/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-bano-lavatorio-peirano-dique-203-acabado-cromado-doble-comando/p/MLA15460849/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/juego-accesorios-bano-peirano-6-piezas-metalicos-3000-color-cromado/p/MLA32486497/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-marbella-30-121-cromo-acabado-cromado/p/MLA8283520/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-vigo-30-132-extensible-acero-inoxidable-acabado-pulido/p/MLA11035200/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/peirano-dique-303-acabado-cromado/p/MLA15460764/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-ducha-peirano-463-dique-embutir-sin-transf/p/MLA21567123/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/peirano-lorca-51-010-blanco-cromado/p/MLA10860513/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-peirano-bedia-51-107-cromo-acabado-cromado/p/MLA9279917/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-peirano-valencia-51-122-acabado-cromado/p/MLA15787476/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-bimando-peirano-linea-mallorca-51-131-acabado-brillante-color-gris/p/MLA22854902/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-peirano-black-velvet-51-190-color-negro/p/MLA15761932/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-cocina-peirano-marbella-black-retractil-para-ollas-acabado-mate-color-negro/p/MLA19859722/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-doble-comando-peirano-lorca-53-010-blanco-y-cromo/p/MLA10948989/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/grifo-de-cocina-monocomando-peirano-bilbao-54-130-cromo-acabado-cromado/p/MLA9273006/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/peirano-lorca-60-010-color-blanca-y-acabado-cromado/p/MLA15462500/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-de-bano-peirano-malba-60-058-color-cromo/p/MLA15443145/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-lavatorio-mesada-peirano-fabric-gold-dorado-cceram/p/MLA22209427/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-de-lavatorio-peirano-fabric-negro-mate-c-ceramico/p/MLA24531432/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-bano-peirano-lavatorio-marbella-cromo-acabado-cromado-color-plateado/p/MLA28707115/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-de-bano-peirano-valencia-60-122-acabado-cromado/p/MLA15846490/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/peirano-lugo-60-133-color-plateada-y-acabado-cromado/p/MLA15452372/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-bano-peirano-lavatorio-santander-cromo-acabado-cromado-color-plateado/p/MLA28155109/s',
#     Producto(codigo ='',link=''codigo = '60-150',link =)'https://www.mercadolibre.com.ar/griferia-lavatorio-toledo-peirano-ceramico/p/MLA22598139/s',
#     Producto(codigo ='',link=''codigo = '60-152',link ='https://www.mercadolibre.com.ar/peirano-soria-60-152-acabado-cromado/p/MLA15842519/s'),
#     Producto(codigo ='',link=''codigo = '60-153',link ='https://www.mercadolibre.com.ar/griferia-de-bano-peirano-castilla-60-153-color-cromo-y-acabado-cromado/p/MLA15831683/s'),
#     Producto(codigo ='',link=''codigo = '60-161',link ='https://www.mercadolibre.com.ar/grifo-lavatorio-bano-peirano-lago-cierre-ceramico-cromo-color-gris/p/MLA26713509/s'),
#     Producto(codigo ='',link=''codigo = '60-162',link ='https://www.mercadolibre.com.ar/griferia-bano-canilla-lavatorio-peirano-adra-vanitory-60-162-acabado-cromado-color-plateado/p/MLA15383205/s'),
#     Producto(codigo ='',link=''codigo = '60-180',link ='https://www.mercadolibre.com.ar/griferia-monocomando-lavatorio-bajo-peirano-vera-60-180-cuot-acabado-cromado-color-plateado/p/MLA15452556/s'),
#     Producto(codigo ='',link=''codigo = '61-180',link ='https://www.mercadolibre.com.ar/peirano-vera-61-180-acabado-cromado/p/MLA15397965/s'),
#     Producto(codigo ='',link=''codigo = '62-100G',link ='https://www.mercadolibre.com.ar/peirano-griferia-fabric-lavatorio-pared-62-100g-gold/p/MLA22430365/s'),
#     Producto(codigo ='',link=''codigo = '62-100N',link ='https://www.mercadolibre.com.ar/griferia-de-pared-peirano-fabric-black-62-100n-lavatorio-acabado-negro-mate-color-negro/p/MLA23552548/s'),
#     Producto(codigo ='',link=''codigo = '62-121',link ='https://www.mercadolibre.com.ar/griferia-peirano-marbella-bano-lavatorio-pared-cromo-62-121-color-gris/p/MLA28193062/s'),
#     Producto(codigo ='',link=''codigo = '62-122',link ='https://www.mercadolibre.com.ar/peirano-valencia-62-122-acabado-cromado/p/MLA15912421/s'),
#     Producto(codigo ='',link=''codigo = '62-162',link ='https://www.mercadolibre.com.ar/griferia-lavatorio-pared-peirano-adra-62-162-cierre-ceramico/p/MLA22353920/s'),
#     Producto(codigo ='',link=''codigo = '70-010',link ='https://www.mercadolibre.com.ar/griferia-bidet-peirano-lorca-70-010-color-blanca-y-acabado-cromado/p/MLA15402659/s'),
#     Producto(codigo ='',link=''codigo = '70-058',link ='https://www.mercadolibre.com.ar/griferia-bidet-peirano-malba-compresion-acabado-cromo-acabado-cromado-color-gris/p/MLA28090361/s'),
#     Producto(codigo ='',link=''codigo = '70-100C',link ='https://www.mercadolibre.com.ar/griferia-peirano-fabric-bidet-cierre-ceramico-70-100c/p/MLA19940510/s'),
#     Producto(codigo ='',link=''codigo = '70-100G',link ='https://www.mercadolibre.com.ar/griferia-de-bidet-peirano-fabric-70-100g-gold/p/MLA21141090/s'),
#     Producto(codigo ='',link=''codigo = '70-100N',link ='https://www.mercadolibre.com.ar/griferia-de-bidet-moderna-peirano-fabric-70-100n-black-redon/p/MLA19727086/s'),
#     Producto(codigo ='',link=''codigo = '70-121',link ='https://www.mercadolibre.com.ar/griferia-de-bano-bidet-peirano-marbella-cromado-70-121-color-plateado/p/MLA28489958/s'),
#     Producto(codigo ='',link=''codigo = '70-122',link ='https://www.mercadolibre.com.ar/peirano-valencia-70-122-acabado-cromado/p/MLA15846262/s'),
#     Producto(codigo ='',link=''codigo = '70-135',link ='https://www.mercadolibre.com.ar/griferia-peirano-santander-bidet-cierre-ceramico-70-135-acabado-cromado-color-cromo/p/MLA29686657/s'),
#     Producto(codigo ='',link=''codigo = '70-152',link ='https://www.mercadolibre.com.ar/peirano-soria-70-152-acabado-cromado/p/MLA16012522/s'),
#     Producto(codigo ='',link=''codigo = '70-153',link ='https://www.mercadolibre.com.ar/peirano-castilla-70-153-color-cromo-y-acabado-cromado/p/MLA15780997/s'),
#     Producto(codigo ='',link=''codigo = '70-161',link ='https://www.mercadolibre.com.ar/griferia-bano-lago-peirano-bidet-p-acabado-cromado-color-gris/p/MLA27943444/s'),
#     Producto(codigo ='',link=''codigo = '70-162',link ='https://www.mercadolibre.com.ar/griferia-canilla-bide-bano-peirano-adra-70-162-acabado-cromado-color-plateado/p/MLA15962905/s'),
#     Producto(codigo ='',link=''codigo = '70-180',link ='https://www.mercadolibre.com.ar/peirano-vera-70-180-acabado-cromado/p/MLA16006750/s'),
#     Producto(codigo ='',link=''codigo = '80-058',link ='https://www.mercadolibre.com.ar/griferia-bano-ducha-peirano-malba-cierrre-compresion/p/MLA25750337/s'),
#     Producto(codigo ='',link=''codigo = '80-100C',link ='https://www.mercadolibre.com.ar/griferia-ducha-fabric-cromo-peirano-80-100c-cc-cr-color-plateado/p/MLA22225458/s'),
#     Producto(codigo ='',link=''codigo = '80-100G',link ='https://www.mercadolibre.com.ar/griferia-bano-ducha-fabric-peirano-dorado-80-100g-acabado-brillante/p/MLA32161697/s'),
#     Producto(codigo ='',link=''codigo = '80-100N',link ='https://www.mercadolibre.com.ar/griferia-bano-ducha-peirano-fabric-80-100n-black-acabado-mate-color-negro/p/MLA31105071/s'),
#     Producto(codigo ='',link=''codigo = '80-122',link ='https://www.mercadolibre.com.ar/peirano-griferia-valencia-ducha-ctransferencia-80-122-acabado-brillante-color-plateado/p/MLA21482872/s'),
#     Producto(codigo ='',link=''codigo = '80-135',link ='https://www.mercadolibre.com.ar/peirano-griferia-santander-ducha-80-135-hot-sale-acabado-cromado-color-plateado/p/MLA22225313/s'),
#     Producto(codigo ='',link=''codigo = '80-136',link ='https://www.mercadolibre.com.ar/griferia-peirano-marbella-ducha-embutir-ctrans-cierre-ceram/p/MLA21617588/s'),
#     Producto(codigo ='',link=''codigo = '80-150',link ='https://www.mercadolibre.com.ar/griferia-peirano-toledo-ducha-ctrans-cierre-ceramico-80-150/p/MLA23229961/s'),
#     Producto(codigo ='',link=''codigo = '80-152',link ='https://www.mercadolibre.com.ar/griferia-ducha-peirano-80-152-soria-embutir-con-transf-acabado-cromado-color-gris/p/MLA28115199/s'),
#     Producto(codigo ='',link=''codigo = '80-153',link ='https://www.mercadolibre.com.ar/griferia-de-ducha-peirano-castilla-80-153-acabado-cromado-color-cromo/p/MLA21788483/s'),
#     Producto(codigo ='',link=''codigo = '80-161',link ='https://www.mercadolibre.com.ar/griferia-peirano-lago-ducha-embutir-con-transferencia-80-161/p/MLA22877018/s'),
#     Producto(codigo ='',link=''codigo = '80-162',link ='https://www.mercadolibre.com.ar/griferia-ducha-peirano-adra-80-162-ceramico-con-transferenci/p/MLA21805332/s'),
#     Producto(codigo ='',link=''codigo = '80-180',link ='https://www.mercadolibre.com.ar/griferia-ducha-80-180-vera-peirano-embutir-con-transferencia/p/MLA21482968/s'),
#     Producto(codigo ='',link=''codigo = '80-352',link ='https://www.mercadolibre.com.ar/griferia-ducha-peirano-embutir-soria-3-funciones-80-352-acabado-cromado-color-plateado/p/MLA21384317/s'),
#     Producto(codigo ='',link=''codigo = '81-010',link ='https://www.mercadolibre.com.ar/griferia-ducha-peirano-81-010-lorca-embutir-sin-transfere-s-acabado-brillante-color-blancocromo/p/MLA21617592/s'),
#     Producto(codigo ='',link=''codigo = '83-010',link ='https://www.mercadolibre.com.ar/griferia-ducha-peirano-83-010-lorca-exterior-ctransf-p/p/MLA21567136/s'),
#     Producto(codigo ='',link=''codigo = 'A813',link ='https://www.mercadolibre.com.ar/flor-de-ducha-3-funciones-con-brazo-peirano-a813/p/MLA22934574/s'),
#     Producto(codigo ='',link=''codigo = 'A814',link ='https://www.mercadolibre.com.ar/flor-de-ducha-peirano-bano-brazo-y-campana-anticalcarea-acabado-cromado-color-plateado/p/MLA22877028/s'),
#     Producto(codigo ='',link=''codigo = 'A815',link ='https://www.mercadolibre.com.ar/flor-de-ducha-peirano-bano-redonda-8-brazo-anticalcareo-acabado-cromado/p/MLA23323241/s'),
#     Producto(codigo ='',link=''codigo = 'A816',link ='https://www.mercadolibre.com.ar/flor-de-ducha-anticalcarea-6-con-brazo-peirano-a816/p/MLA22988814/s'),
#     Producto(codigo ='',link=''codigo = 'BCH03/3',link = 'https://www.mercadolibre.com.ar/bacha-peirano-cuadrada-apoyo-3-agujeros-5-anos-50-x-42-cm/p/MLA24547878/s'),
#     Producto(codigo ='',link=''codigo = 'BCH04/3',link = 'https://www.mercadolibre.com.ar/bacha-apoyar-cuadrada-3aug-420x420x125-peirano-bch043-acabado-blanco-color-blanco/p/MLA23317041/s')
# ]



def clean_normalize_name(name):
    cleaned_name = name.replace(",", "").replace("°", "").replace(" ", "_").replace("*"," ").upper()
    normalized_name = unicodedata.normalize('NFD', cleaned_name)
    normalized_name = ''.join(c for c in normalized_name if not unicodedata.combining(c))

    return normalized_name


def get_xpath_or_unavailable(driver,xpath):
    try:
        element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, xpath)))
        return element.text.strip() if element.text else 'No disponible'
    except Exception as e:
        return 'No disponible'


def buscar_competidor(driver,competidor,lista_producto):
    precio = get_xpath_or_unavailable(driver,competidor[0])
    print(f"Precio es {precio}")
    lista_producto.append(precio)

    vendedor = get_xpath_or_unavailable(driver,competidor[1])
    print(f"Vendedor es {vendedor}")
    lista_producto.append(vendedor)


def run_program(producto,driver):

    print(f"\n Codigo es {producto.codigo}")

    lista_producto = [producto.codigo]

    # Navegar a la URL deseada
    driver.get(producto.link)
    # Extraer el nombre
    nombre = get_xpath_or_unavailable(driver,producto.nombre)

    print(f"Nombre es {nombre}")

    lista_producto.append(nombre)
    # Esperar a que aparezca el elemento con el precio deseado y extraer el precio

    buscar_competidor(driver,producto.primero,lista_producto)
    buscar_competidor(driver,producto.segundo,lista_producto)

    return lista_producto




def escribirExcel(lista_a_cargar,hoja_destino,excel):
    # Cargar el archivo Excel

    print("Estoy en la hoja")
    
    for row in excel[hoja_destino].iter_rows(min_row=3, max_row=779, min_col=1, max_col=1):
        for producto_cargado in lista_a_cargar:
            if producto_cargado[0] == (row[0].value):
                print("Coincidencia encontrada. Sobrescribiendo...")
                # Sobrescribir los datos en la fila correspondiente
                row[0].offset(column=1).value = producto_cargado[1]
                row[0].offset(column=2).value = producto_cargado[2]
                row[0].offset(column=3).value = producto_cargado[3]
                row[0].offset(column=4).value = producto_cargado[4]
                row[0].offset(column=5).value = producto_cargado[5]
                row[0].offset(column=6).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                break
    
    # Guardar el archivo Excel
    excel.save(ruta_excel)

# def obtener_urls(columna, hasta_fila):
#     workbook = load_workbook(ruta_excel)
#     sheet = workbook['Fv']
    
#     urls_fv = []
#     for row in sheet.iter_rows(min_row=1, max_row=hasta_fila, min_col=columna, max_col=columna):
#         for cell in row:
#             if(cell.value == 'Cata'):
#                 if cell.hyperlink:
#                     url = cell.hyperlink.target
#                     urls_fv.append(url)
    
#     workbook.close()  # No olvides cerrar el archivo después de leerlo
    
#     return urls_fv 



def cargar_lista_productos(lista_productos,driver,lista_a_cargar = []):
    for producto in lista_productos:
        lista_producto = run_program(producto,driver)
        lista_a_cargar.append(lista_producto)
    return lista_a_cargar


def ejectar_programa():
    print("Empezamos a laburaaaar")

    excel = SeniorExcel(excel= load_workbook(ruta_excel))
    lista_productos_fv = excel.cargar_productos('FvCata')
    lista_productos_ferrum = excel.cargar_productos('FerrumCata')
    lista_productos_peirano = excel.cargar_productos('Peirano')
    # urls_fv = obtener_urls(4, 170)
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # Inicializar el driver de Chrome
    driver = webdriver.Chrome(options=chrome_options)

    print("Entro a chrome")

    lista_a_cargar_fv = cargar_lista_productos(lista_productos_fv,driver)
    lista_a_cargar_ferrum = cargar_lista_productos(lista_productos_ferrum,driver)
    lista_a_cargar_peirano = cargar_lista_productos(lista_productos_peirano,driver)

    print("voy a escribir")
    
    print("Escribo FVCata")    
    excel.escribirExcel(lista_a_cargar_fv,'FvCata')
    print("Escribo FerrumCata")  
    excel.escribirExcel(lista_a_cargar_ferrum,'FerrumCata')
    print("Escribo Peirano")  
    excel.escribirExcel(lista_a_cargar_peirano,'Peirano')

    
    

try:
    ejectar_programa()
except:
    print("\nEl archivo de Excel 'FvCata.xlsx' no se encuentra en la misma carpeta que este programa o se ha borrado.")
    print("Vuelva a insertarlo para que el programa funcione correctamente.")
